"""
Consolida todos los archivos Excel mensuales de la CMF en dos DataFrames:
- cuadro1: índice de provisiones por banco y tipo de crédito
- cuadro2: calidad de cartera (normal, subestándar, en incumplimiento)
"""

import re
from pathlib import Path

import openpyxl
import pandas as pd
import xlrd

RAW_DIR = Path(__file__).parent.parent / "data" / "raw"
OUT_DIR = Path(__file__).parent.parent / "data" / "processed"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Bancos que no son instituciones reales (filas de totales o separadores)
FILAS_IGNORAR = {"sistema bancario", "total"}


def fecha_desde_nombre(path: Path) -> pd.Timestamp | None:
    """Extrae año y mes desde el nombre del archivo (ej: cmf_2024_05.xlsx)."""
    m = re.search(r"cmf_(\d{4})_(\d{2})", path.stem)
    if not m:
        return None
    return pd.Timestamp(year=int(m.group(1)), month=int(m.group(2)), day=1)


def limpiar_nombre_banco(nombre: str) -> str:
    """Saca asteriscos y notas al pie de los nombres de bancos.
    Ej: 'Banco de Chile (**)' -> 'Banco de Chile'
    """
    return re.sub(r"\s*\([\*\d]+\)", "", nombre).strip()


def es_banco(valor) -> bool:
    """Filtra filas que no son bancos: totales, nulos o separadores."""
    if not isinstance(valor, str) or not valor.strip():
        return False
    limpio = limpiar_nombre_banco(valor).lower()
    return limpio not in FILAS_IGNORAR and not limpio.startswith("nota")


def extraer_cuadro1(ws, fecha: pd.Timestamp) -> list[dict]:
    """
    Cuadro 1: índice de provisiones por banco, desglosado en total,
    comercial, consumo y vivienda. Los datos siempre arrancan en fila 15.
    """
    registros = []

    for row in ws.iter_rows(min_row=15, values_only=True):
        banco = row[1] if len(row) > 1 else None
        if not es_banco(banco):
            continue

        # Si ya llegamos a las notas al pie, paramos
        if isinstance(banco, str) and banco.strip().startswith("("):
            break

        registros.append({
            "fecha": fecha,
            "banco": limpiar_nombre_banco(banco),
            "indice_total": row[2] if len(row) > 2 else None,
            "indice_comercial": row[3] if len(row) > 3 else None,
            "indice_consumo": row[4] if len(row) > 4 else None,
            "indice_vivienda": row[5] if len(row) > 5 else None,
        })

    return registros


def extraer_cuadro2(ws, fecha: pd.Timestamp) -> list[dict]:
    """
    Cuadro 2: calidad de cartera — índice de provisiones según evaluación
    individual/grupal y clasificación (normal, subestándar, en incumplimiento).
    Los datos arrancan en fila 14.
    """
    registros = []

    for row in ws.iter_rows(min_row=14, values_only=True):
        banco = row[1] if len(row) > 1 else None
        if not es_banco(banco):
            continue

        if isinstance(banco, str) and banco.strip().startswith("("):
            break

        registros.append({
            "fecha": fecha,
            "banco": limpiar_nombre_banco(banco),
            "indice_ev_individual": row[2] if len(row) > 2 else None,
            "indice_ev_grupal": row[4] if len(row) > 4 else None,
            "indice_cartera_normal": row[6] if len(row) > 6 else None,
            "indice_cartera_subestandar": row[8] if len(row) > 8 else None,
            "indice_cartera_incumplimiento": row[10] if len(row) > 10 else None,
        })

    return registros


def procesar_archivo(path: Path) -> tuple[list, list]:
    """Lee un archivo Excel y devuelve los registros de cuadro 1 y 2."""
    fecha = fecha_desde_nombre(path)
    if not fecha:
        print(f"  Saltando {path.name} — no se pudo leer la fecha del nombre")
        return [], []

    # Los .xls antiguos usan xlrd, los .xlsx usan openpyxl
    if path.suffix.lower() == ".xls":
        return procesar_xls(path, fecha)

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"  Error leyendo {path.name}: {e}")
        return [], []

    # Los nombres de las hojas tienen problemas de encoding, usamos índice
    sheet_names = wb.sheetnames
    c1 = extraer_cuadro1(wb[sheet_names[1]], fecha) if len(sheet_names) > 1 else []
    c2 = extraer_cuadro2(wb[sheet_names[2]], fecha) if len(sheet_names) > 2 else []

    wb.close()
    return c1, c2


def procesar_xls(path: Path, fecha: pd.Timestamp) -> tuple[list, list]:
    """Procesa archivos .xls (formato antiguo) usando xlrd."""
    try:
        wb = xlrd.open_workbook(path)
    except Exception as e:
        print(f"  Error leyendo {path.name}: {e}")
        return [], []

    def iter_rows(sheet, min_row=0):
        for i in range(min_row, sheet.nrows):
            yield sheet.row_values(i)

    # Cuadro 1 es la hoja índice 1, Cuadro 2 es la hoja índice 2
    c1, c2 = [], []

    if wb.nsheets > 1:
        ws1 = wb.sheet_by_index(1)
        for row in iter_rows(ws1, min_row=14):
            banco = row[1] if len(row) > 1 else None
            if not es_banco(banco):
                continue
            if isinstance(banco, str) and banco.strip().startswith("("):
                break
            c1.append({
                "fecha": fecha,
                "banco": limpiar_nombre_banco(banco),
                "indice_total": row[2] if len(row) > 2 else None,
                "indice_comercial": row[3] if len(row) > 3 else None,
                "indice_consumo": row[4] if len(row) > 4 else None,
                "indice_vivienda": row[5] if len(row) > 5 else None,
            })

    if wb.nsheets > 2:
        ws2 = wb.sheet_by_index(2)
        for row in iter_rows(ws2, min_row=13):
            banco = row[1] if len(row) > 1 else None
            if not es_banco(banco):
                continue
            if isinstance(banco, str) and banco.strip().startswith("("):
                break
            c2.append({
                "fecha": fecha,
                "banco": limpiar_nombre_banco(banco),
                "indice_ev_individual": row[2] if len(row) > 2 else None,
                "indice_ev_grupal": row[4] if len(row) > 4 else None,
                "indice_cartera_normal": row[6] if len(row) > 6 else None,
                "indice_cartera_subestandar": row[8] if len(row) > 8 else None,
                "indice_cartera_incumplimiento": row[10] if len(row) > 10 else None,
            })

    return c1, c2


def limpiar_df(df: pd.DataFrame) -> pd.DataFrame:
    """Reemplaza '---' por NaN y convierte columnas numéricas."""
    df = df.replace("---", pd.NA)
    for col in df.columns:
        if col not in ("fecha", "banco"):
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def main():
    archivos = sorted(RAW_DIR.glob("cmf_*.xl*"))
    print(f"Archivos encontrados: {len(archivos)}\n")

    todos_c1, todos_c2 = [], []

    for i, path in enumerate(archivos, 1):
        print(f"[{i:>3}/{len(archivos)}] {path.name}")
        c1, c2 = procesar_archivo(path)
        todos_c1.extend(c1)
        todos_c2.extend(c2)

    # Armar y limpiar DataFrames
    df_c1 = limpiar_df(pd.DataFrame(todos_c1))
    df_c2 = limpiar_df(pd.DataFrame(todos_c2))

    # Ordenar por fecha y banco
    df_c1 = df_c1.sort_values(["fecha", "banco"]).reset_index(drop=True)
    df_c2 = df_c2.sort_values(["fecha", "banco"]).reset_index(drop=True)

    # Guardar
    df_c1.to_csv(OUT_DIR / "provisiones_por_tipo.csv", index=False)
    df_c2.to_csv(OUT_DIR / "calidad_cartera.csv", index=False)

    print(f"\nDataset cuadro 1: {df_c1.shape[0]} filas × {df_c1.shape[1]} columnas")
    print(f"Dataset cuadro 2: {df_c2.shape[0]} filas × {df_c2.shape[1]} columnas")
    print(f"\nBancos encontrados: {sorted(df_c1['banco'].unique())}")
    print(f"\nRango de fechas: {df_c1['fecha'].min().date()} a {df_c1['fecha'].max().date()}")
    print("\nArchivos guardados en data/processed/")


if __name__ == "__main__":
    main()
